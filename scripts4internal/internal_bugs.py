import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# List of Internal bugs which are created by Manjesh N
manjeshn_reported_bugs = ["NSUI-19176", "NSUI-19185", "NSUI-19186", "NSUI-19187", "NSUI-19188", "NSUI-19190", "NSAUTH-13672", "NSUI-19215", "NSUI-19214", "NSUI-19892", "NSLINUX-7296", "NSLINUX-7297", "NSLINUX-7298", "NSADM-111975", "NSLINUX-7299", 
                          "NSLINUX-7300", "NSLINUX-7301", "NSADM-111979", "NSLINUX-7303", "NSLINUX-7304", "NSLINUX-7305", "NSADM-112040", "NSLINUX-7306", "NSLINUX-7307", "NSADM-112298", "NSLINUX-7338", "NSLINUX-7345", "NSLINUX-7346", "NSADM-112541",
                          "NSADM-112685", "NSDOC-3777", "NSUI-19962", "NSUI-19963", "NSUI-19964", "NSUI-19965", "NSADM-112734", "NSADM-112738", "NSDOC-3787", "NSADM-113020", "NSADM-113022", "NSADM-113039", "NSADM-113125", "NSADM-113126", "NSADM-113127",
                          "NSADM-113128", "NSADM-113175", "NSADM-113176", "NSADM-113281", "NSADM-113316", "NSADM-113317", "NSADM-113318", "NSLINUX-7403", "NSUI-20009", "NSUI-20010", "NSUI-20011", "NSUI-20012", "NSUI-20013", "NSUI-20014", "NSUI-20015",
                          "NSUI-20016", "NSADM-113500", "NSADM-113996"]

# Required URL struct
jira_url = 'https://issues.citrite.net/rest/api/2/issue/'

# Authorization tokens
headers = {"Authorization": "Bearer MzAxNTE5MjE5MzU0OtUTcabB1CEMVwUGotGeibLeghra"}

# Create a workbook and select active sheet
wb = Workbook()
ws = wb.active

# Rename the active sheet to "Internal_Bugs"
ws.title = "Internal_Bugs"

# Write headers
header_row = ["JIRA ID", "Project", "Severity", "Summary", "Status", "Resolution", "Jira Owner", "Created Date", "Resolution Date", "TTR", "Reporter"]
ws.append(header_row)

# Highlight header row
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

# Dictionaries to store project and status counts
project_counts = {}
status_counts = {}

# Fetch data for each bug
for bug in manjeshn_reported_bugs:
    bug_url = f'{jira_url}{bug}'
    response = requests.get(bug_url, headers=headers)
    if response.status_code == 200:
        jsonData = response.json()
        jira_id = jsonData['key']
        summary = jsonData['fields']['summary']
        project = jsonData['fields']['project']['key']
        status = jsonData['fields']['status']['name']
        resolution = jsonData['fields']['resolution']['name'] if jsonData['fields'].get('resolution') is not None and 'name' in jsonData['fields']['resolution'] else 'Unresolved'
        raw_resolutiondate = jsonData['fields']['resolutiondate']
        resolutiondate = ''
        if raw_resolutiondate is not None:
            resolutiondate = datetime.strptime(raw_resolutiondate, "%Y-%m-%dT%H:%M:%S.%f%z")
        created_date = datetime.strptime(jsonData['fields']['created'], "%Y-%m-%dT%H:%M:%S.%f%z")
        severity = jsonData['fields']['customfield_18130']['value']
        reporter = jsonData['fields']['reporter']['displayName']
        assignee_name = jsonData['fields']['assignee']['displayName'] if jsonData['fields'].get('assignee') is not None else 'Unassigned'
        duration = '' if resolutiondate == '' else (resolutiondate - created_date).days

        # Append the data to the worksheet
        row = [jira_id, project, severity, summary, status, resolution, assignee_name, created_date.strftime("%d-%b-%Y"), resolutiondate.strftime("%d-%b-%Y") if resolutiondate else '', duration, reporter]
        ws.append(row)
        
        # Increment the project count
        if project in project_counts:
            project_counts[project] += 1
        else:
            project_counts[project] = 1
        
        # Increment the status count
        if status in status_counts:
            status_counts[status] += 1
        else:
            status_counts[status] = 1

        # Change the cell background color based on status
        status_cell = ws.cell(row=ws.max_row, column=5)
        if status == "Done":
            status_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        elif status == "Canceled":
            status_cell.fill = PatternFill(start_color="F83C3C", end_color="F83C3C", fill_type="solid")  # Red
        elif status == "In Progress":
            status_cell.fill = PatternFill(start_color="1AC2F3", end_color="1AC2F3", fill_type="solid")  # Blue
        elif status == "Backlog":
            status_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        elif status == "Ready For Test":
            status_cell.fill = PatternFill(start_color="F88C3C", end_color="F88C3C", fill_type="solid")  # Orange
        elif status == "Planned":
            status_cell.fill = PatternFill(start_color="ED1CF7", end_color="ED1CF7", fill_type="solid")  # Magenta
        elif status == "In Review":
            status_cell.fill = PatternFill(start_color="40FDB2", end_color="40FDB2", fill_type="solid")  # Green Kind

# Auto-adjust column width
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

# Create BreakUp_Details sheet
breakup_ws = wb.create_sheet(title="BreakUp_Details")

# Write headers for BreakUp_Details
breakup_header_row = ["Project", "Count"]
breakup_ws.append(breakup_header_row)

# Highlight header row for BreakUp_Details
for cell in breakup_ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

# Populate BreakUp_Details sheet with project counts
for project, count in project_counts.items():
    breakup_ws.append([project, count])

# Add grand total row
grand_total = sum(project_counts.values())
breakup_ws.append(["Grand Total", grand_total])
grand_total_cell = breakup_ws.cell(row=breakup_ws.max_row, column=2)
grand_total_cell.font = Font(bold=True)
grand_total_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow

# Add borders to all cells in BreakUp_Details
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in breakup_ws.iter_rows(min_row=1, max_row=breakup_ws.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.border = thin_border

# Auto-adjust column width for BreakUp_Details
for column in breakup_ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    breakup_ws.column_dimensions[column_letter].width = adjusted_width

# Add Status breakup details
breakup_ws.append([])  # Blank row for separation
breakup_ws.append(["Status", "Count"])  # Header for status breakup
for cell in breakup_ws[breakup_ws.max_row]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")

# Populate BreakUp_Details sheet with status counts
for status, count in status_counts.items():
    breakup_ws.append([status, count])

# Add borders to the status table cells
for row in breakup_ws.iter_rows(min_row=breakup_ws.max_row-len(status_counts), max_row=breakup_ws.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.border = thin_border

# Auto-adjust column width for status breakup
for column in breakup_ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    breakup_ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
wb.save(r"G:\My Drive\bug_report.xlsx")
